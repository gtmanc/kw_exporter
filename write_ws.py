from openpyxl import load_workbook
from openpyxl import Workbook

"""
Write cells in first row, in default sheet with specified names.
list: names will be written in the cells in first row
"""
def init(list):

    wb = Workbook()
    #wb = load_workbook('mainbuilding33.xlsx')
    ws = wb.active #get a defult sheet
    #print(ws.title)
    # ws.title = "first" #give a name to a sheet
    # ws2 = wb.create_sheet("my_sheet") #create a new sheet
    # ws = wb["first"] #get a heet with a specified name
    
    for i in range(len(list)):
        ws.cell(row = 1, column = i+1, value = list[i])

    return wb

"""
Create a work book.
Input
None

output:
work book
"""
def create_wb():
  wb = Workbook()
  return wb


"""
Create a work sheet and fill 1st row with names
Input
name: worksheet name. If the length of the name is longer than 31, it is cut because excel warns. 
list: name to be filled

Output
Work sheet with 1st row filled
"""
def create_ws(wb, name, list):
  if len(name) > 31:
    name = name[:31]

  ws = wb.create_sheet(name, 0) #Always put the sheet to the head

  for i in range(len(list)):
        ws.cell(row = 1, column = i+1, value = list[i])

  #Change background color
  #ws.sheet_properties_tabColor = '1072BA'

  return ws

"""
Write cells in a row starting with first cell. 
list: values to be wriiten to the the cell in the row
ws: worksheet
row_num: row number in the worksheet which will be written
"""
def write2ws(list, ws, row_num):
    if row_num == 0:
        print('invalid row number')
        return

    for i in range(len(list)):
      #print('Item = {t}'.format(t = list[i]))
      ws.cell(row = row_num, column = i+1, value = list[i])

    return row_num